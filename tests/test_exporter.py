import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import io
from datetime import date
import pytest
from openpyxl import load_workbook
from exporter import generate_excel
from vaccine_logic import APLICADA, PENDIENTE


def _make_patient(has_pending=False):
    vaccines = {
        "BCG": [{"seq": 1, "label": "DU", "status": APLICADA,
                 "applied_date": date(2024, 12, 29), "applied_label": "DU"}],
        "HVB": [{"seq": 1, "label": "DU",
                 "status": PENDIENTE if has_pending else APLICADA,
                 "applied_date": None if has_pending else date(2024, 12, 29),
                 "applied_label": None if has_pending else "DU"}],
    }
    for col in ["PENTAVALENTE", "IPV", "ROTAVIRUS", "NEUMOCOCO",
                "INFLUENZA PEDIATRICA", "INFLUENZA ADULTO", "SPR", "VARICELA",
                "HEPATITIS A", "AMARILICA", "DPT", "APO", "dT", "HiB"]:
        vaccines[col] = []
    return {
        "DNI": "12345678", "Nombres": "TEST PACIENTE", "Sexo": "MASCULINO",
        "F_Nacimiento": "01/01/2025", "Edad": "1a 3m 9d", "Grupo": "1 año",
        "Red": "RED A", "Microred": "MICRO A", "EESS": "EESS A",
        "vaccines": vaccines,
    }


def test_generate_excel_returns_bytes():
    result = generate_excel([_make_patient(has_pending=False)])
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_excel_has_two_sheets():
    result = generate_excel([_make_patient(True), _make_patient(False)])
    wb = load_workbook(io.BytesIO(result))
    assert "Completo" in wb.sheetnames
    assert "Pendientes" in wb.sheetnames


def test_pendientes_sheet_only_has_pending_patients():
    result = generate_excel([_make_patient(True), _make_patient(False)])
    wb = load_workbook(io.BytesIO(result))
    ws = wb["Pendientes"]
    assert ws.max_row == 2   # header + 1 pending patient


def test_completo_sheet_has_all_patients():
    result = generate_excel([_make_patient(True), _make_patient(False)])
    wb = load_workbook(io.BytesIO(result))
    ws = wb["Completo"]
    assert ws.max_row == 3   # header + 2 patients


def test_pendientes_sheet_has_extra_column():
    result = generate_excel([_make_patient(True)])
    wb = load_workbook(io.BytesIO(result))
    ws = wb["Pendientes"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "Vacunas pendientes" in headers
