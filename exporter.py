# -*- coding: utf-8 -*-
"""Generador del reporte Excel con 2 hojas y colores por estado vacunal."""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from parser import VACCINE_COLUMNS
from vaccine_logic import (
    worst_status_color, patient_has_pending, patient_pending_list,
    format_dose_cell,
)

INFO_HEADERS = [
    "DNI", "Nombres", "Sexo", "F.Nacimiento", "Edad",
    "Grupo", "Red", "Microred", "EESS",
]
ALL_HEADERS = INFO_HEADERS + VACCINE_COLUMNS


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _write_header(ws, headers: list) -> None:
    header_fill = _fill("1A6FA8")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_patient_row(ws, row_idx: int, patient: dict, extra_cols: list | None = None) -> None:
    info_values = [
        patient["DNI"], patient["Nombres"], patient["Sexo"],
        patient["F_Nacimiento"], patient["Edad"], patient["Grupo"],
        patient["Red"], patient["Microred"], patient["EESS"],
    ]
    for col_idx, val in enumerate(info_values, 1):
        ws.cell(row=row_idx, column=col_idx, value=val).alignment = Alignment(wrap_text=False)

    offset = len(INFO_HEADERS) + 1
    for col_idx, vcol in enumerate(VACCINE_COLUMNS, offset):
        dose_results = patient["vaccines"].get(vcol, [])
        text  = format_dose_cell(dose_results)
        color = worst_status_color(dose_results)
        cell  = ws.cell(row=row_idx, column=col_idx, value=text)
        cell.fill      = _fill(color)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if extra_cols:
        extra_offset = offset + len(VACCINE_COLUMNS)
        for col_idx, val in enumerate(extra_cols, extra_offset):
            ws.cell(row=row_idx, column=col_idx, value=val)


def _autowidth(ws, max_width: int = 40) -> None:
    for col in ws.columns:
        # Para celdas multi-línea, medir la línea más larga (no el total)
        length = max(
            (
                max((len(line) for line in str(cell.value or "").split("\n")), default=0)
                for cell in col
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def generate_excel(processed_patients: list) -> bytes:
    """
    Genera el reporte Excel con 2 hojas.

    Hoja 1 - Completo: todos los pacientes con estado de cada dosis.
    Hoja 2 - Pendientes: solo pacientes con dosis PENDIENTE o ERROR_DU,
             con columna extra "Vacunas pendientes".

    Retorna bytes del archivo .xlsx.
    """
    wb = Workbook()

    # ── Hoja 1: Completo ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Completo"
    _write_header(ws1, ALL_HEADERS)
    for i, patient in enumerate(processed_patients, 2):
        _write_patient_row(ws1, i, patient)
    ws1.freeze_panes = "A2"
    _autowidth(ws1)

    # ── Hoja 2: Pendientes ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pendientes")
    _write_header(ws2, ALL_HEADERS + ["Vacunas pendientes"])
    row2 = 2
    for patient in processed_patients:
        if patient_has_pending(patient["vaccines"]):
            pending_str = patient_pending_list(patient["vaccines"])
            _write_patient_row(ws2, row2, patient, extra_cols=[pending_str])
            row2 += 1
    ws2.freeze_panes = "A2"
    _autowidth(ws2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
